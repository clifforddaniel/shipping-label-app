import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from pathlib import Path
import openpyxl
import re
import sys
from collections import defaultdict


# === GUI Setup ===
window = tk.Tk()
window.title("Shipping Label Generator")



# === Paths and State Variables ===
base_path = Path(sys.executable).parent if getattr(sys, 'frozen', False) else Path(__file__).parent

source_folder_path = ""
destination_folder_path = ""
source_mode = tk.StringVar(value="folder")  # default


# === Template and UI Variables ===
template_var = tk.StringVar(value="Select Template")
template1_color_var = tk.StringVar()
template3_color_var = tk.StringVar()
template3_style_var = tk.StringVar()
overwrite_all = None
store_ready_var = tk.BooleanVar(value=False)
pre_ticketed_var = tk.BooleanVar(value=False)
auto_style_var = tk.BooleanVar()

# === Dynamic Style Metadata Storage ===
style_metadata = {}
style_fields = {}

# === Constants == 
SIZES = ["XS", "S", "M", "L", "XL", "2XL", "3XL", "4XL"]

# === Helper Functions ===
def get_size_ratio_string(carton):
    paired = [(label, qty or 0) for label, qty in zip(SIZES, carton["size_quantities"]) if qty]
    if not paired:
        return ("", "")
    
    if template_var.get() == "Template 2":
        ratio_string = ", ".join(f"{label} ({qty})" for label, qty in paired)
        return (ratio_string, "")  # qty string not needed in this format
    else:
        ratio_string = "/".join(label for label, _ in paired)
        qty_string = "/".join(str(qty) for _, qty in paired)
        return ratio_string, qty_string

def is_valid_path(src, dest):
    if not dest or not src:
        messagebox.showerror("Path Not Set", "Please select a source and destination folder before generating labels.")
        return False
    elif dest == "" or src == "":
        messagebox.showerror("Invalid Path", "Please try again. Do not exit the file dialog box without choosing a path or \"cancel\".")
        return False
    else:
        return True

def confirm_overwrite_if_needed(out_path):
    global overwrite_all

    if not out_path.exists() or overwrite_all is True:
        return True

    if overwrite_all is False:
        return False

    dialog = tk.Toplevel(window)
    dialog.title("Overwrite Confirmation")
    tk.Label(dialog, text=f"'{out_path.name}' already exists.\nDo you want to overwrite it?").pack(padx=20, pady=10)

    response = {"choice": None}

    def choose(option):
        response["choice"] = option
        dialog.destroy()

    btn_frame = tk.Frame(dialog)
    btn_frame.pack(pady=10)

    tk.Button(btn_frame, text="Yes", width=10, command=lambda: choose("yes")).pack(side="left", padx=5)
    tk.Button(btn_frame, text="Yes to All", width=10, command=lambda: choose("yes_all")).pack(side="left", padx=5)
    tk.Button(btn_frame, text="No", width=10, command=lambda: choose("no")).pack(side="left", padx=5)

    dialog.grab_set()
    window.wait_window(dialog)

    if response["choice"] == "yes_all":
        overwrite_all = True
        return True
    elif response["choice"] == "yes":
        return True
    elif response["choice"] == "no":
        return False

def get_input_files(source_path):
    source = Path(source_path)
    if source.is_file():
        return [source]  # Just one file
    elif source.is_dir():
        return list(source.glob("*.xlsx"))  # All Excel files in folder
    else:
        return []
    
def collect_unique_styles():
    if not source_folder_path:
        return {}

    from openpyxl import load_workbook
    path = Path(source_folder_path)
    if not path.exists():
        return {}

    file_list = [path] if path.is_file() else list(path.glob("*.xlsx"))
    styles_by_file = {}

    for file in file_list:
        filename = file.name
        wb = load_workbook(file, read_only=True, data_only=True)
        ws = wb.active

        for row in ws.iter_rows(min_row=17, values_only=True):
            non_empty = [cell for idx, cell in enumerate(row[:6]) if idx not in (3, 5)]
            if all(cell is None for cell in non_empty):
                break

            style, desc = row[8], row[9]
            if style and desc:
                styles_by_file.setdefault(filename, set()).add((style, desc))

    # Convert sets to sorted lists for display
    return {file: sorted(styles) for file, styles in styles_by_file.items()}



def sync_style_metadata():
    if auto_style_var.get():
        for key, fields in style_fields.items():
            style_metadata[key] = {k: v.get().strip() for k, v in fields.items()}

def update_auto_style_visibility():
    valid_paths = bool(source_folder_path) and bool(destination_folder_path)
    selected_template = template_var.get() in ["Template 1", "Template 3"]

    if valid_paths and selected_template:
        auto_style_frame.pack(pady=(5, 0))
    else:
        auto_style_frame.pack_forget()


# === Path selection ===
def choose_source():# Source can be one file or a folder
    global source_folder_path
    if source_mode.get() == "file":
        path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
    else:
        path = filedialog.askdirectory()

    source_folder_path = path
    print("Source:", source_folder_path)
    source_label.config(text=source_folder_path)
    update_auto_style_visibility()

def choose_destination_folder():
    global destination_folder_path
    destination_folder_path = filedialog.askdirectory()
    print("Destination folder:", destination_folder_path)
    destination_label.config(text=destination_folder_path)
    update_auto_style_visibility()


# === Parsing Logic ===
def parse_packing_header(ws):

    """
    Extracts general shipping/invoice data from the top region of the packing list.
    """

    header_data = {

        "ship_to_address_line1": ws["B5"].value,
        "ship_to_address_line2": ws["B6"].value,
        "ship_to_address_line3": ws["B7"].value,
        "ship_to_address_line4": ws["B8"].value,

        "shipper_address_line1": ws["L5"].value,
        "shipper_address_line2": ws["L6"].value,
        "shipper_address_line3": ws["L7"].value,

        "invoice_number": ws["H10"].value,

        "total_units": ws["S14"].value,
        "total_weight": round(ws["I14"].value, 1),

        "cubic_feet": round(ws["C14"].value, 1),
        # Add more as needed

    }

    # ===Error handling the PO Box and pallets variables, since the formatting may be inconsistent

    # === PO Box handling ===
    primary_po_cell = ws["C10"].value
    fallback_po_cell = ws["B10"].value

    if primary_po_cell is None:
        match = re.search(r"PO#:\s*([\d\s]+)", str(fallback_po_cell))
        if match:
            header_data["po_box"] = match.group(1).strip()
        else:
            header_data["po_box"] = str(fallback_po_cell).strip()
    else:
        header_data["po_box"] = str(primary_po_cell).strip()

    # === Pallet number handling ===
    primary_pallet_cell = ws["C12"].value
    fallback_pallet_cell = ws["B12"].value

    if primary_pallet_cell is None:
        match = re.search(r"# of Pallets:\s*([\d\s]+)", str(fallback_pallet_cell))
        if match:
            header_data["num_of_pallets"] = match.group(1).strip()
        else:
            header_data["num_of_pallets"] = None
    else:
        value = str(primary_pallet_cell).strip()
        header_data["num_of_pallets"] = value if value and value != "# of Pallets:" else None


        
    return header_data

# Parse the packing list
def parse_packing_list(ws, start_row=17):
    cartons = []

    for row in ws.iter_rows(min_row=start_row, values_only=True):
        # Stop parsing when rows are empty except columns D and F (which sometimes are filled in otherwise empty rows)
        non_empty = [cell for idx, cell in enumerate(row[:6]) if idx not in (3, 5)]
        if all(cell is None for cell in row[:6]):
            break

        carton = {
            "carton_number": row[1],
            "carton_dimension1": row[2],
            "carton_dimension2": row[4],
            "carton_dimension3": row[6],
            "weight": row[7],
            "vendor_style": row[8],
            "description": row[9],
            "size_quantities": row[10:18],
            "total_units": row[18],
        }

        cartons.append(carton)
    return cartons


# === Label Generation Functions === 
def generate_labels():
    global overwrite_all
    overwrite_all = None

    selected = template_var.get()
    if selected == "Template 1":
        generate_template1_labels()
    elif selected == "Template 2":
        generate_template2_labels()
    elif selected == "Template 3":
        generate_template3_labels()
    else:
        messagebox.showwarning("No Template Selected", "Please choose a template.")

# Label templates
def generate_template1_labels():
    if not is_valid_path(source_folder_path, destination_folder_path):
        return
    
    print("Running Template 1 label logic...")
    print("From:", source_folder_path)
    print("To:", destination_folder_path)

    # Grab user inputs
    sync_style_metadata()

    saved_count = 0
    for file in get_input_files(source_folder_path): # Loop xlsx files in destination folder
        if file.name.startswith("~$"):
            print("Skipping temporary file:", file.name) #Skip temporary files created by Excel
            continue

        out_path = Path(destination_folder_path) / f"{file.stem}-LABELS.xlsx"

        if not confirm_overwrite_if_needed(out_path):
            print("Skipped:", out_path.name)
            continue
        
        print("Processing ", file.name)
    
        # Load input packing list
        source_wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
        packing_list = source_wb.active

        # Load template workbook
        template_path = base_path / "templates" / "template1.xlsx"
        label_wb = openpyxl.load_workbook(template_path, data_only=True)
        template = label_wb.active

        # === Parse and display header info ===
        header = parse_packing_header(packing_list)
        print("Header data:", header)

        # === Parse and display carton info ===
        cartons = parse_packing_list(packing_list)
        print("Carton data: ", cartons)


        # === Output to label template (stub for now) ===
        store_ready = "Yes" if store_ready_var.get() else "No"
        pre_ticketed = "Yes" if pre_ticketed_var.get() else "No"
        for i, carton in enumerate(cartons, start=1):
            print(f"Carton {i} of {len(cartons)}")

            key = (file.name, carton["vendor_style"], carton["description"])
            meta = style_metadata.get(key)
            color = meta["color"] if meta else template1_color_var.get().strip()

            
            new_sheet = label_wb.copy_worksheet(template)
            new_sheet.title = f"Carton {i}"
            # TODO: Manually map carton + header values to template cells
            ratio, qtys = get_size_ratio_string(carton)

            new_sheet["G4"] = header["ship_to_address_line1"]
            new_sheet["G5"] = header["ship_to_address_line2"]
            new_sheet["G6"] = header["ship_to_address_line3"]
            new_sheet["G7"] = header["ship_to_address_line4"]

            new_sheet["C4"] = header["shipper_address_line1"]
            new_sheet["C5"] = f'{header["shipper_address_line2"]}, {header["shipper_address_line3"]}'
            new_sheet["C7"] = header["po_box"]
            new_sheet["E11"] = ratio 
            new_sheet["E12"] = qtys
            new_sheet["B11"] = f'{carton["description"]} # {carton["vendor_style"]}'
            new_sheet["G11"] = color
            new_sheet["I11"] = carton["total_units"]
            new_sheet["C14"] = store_ready
            new_sheet["C15"] = pre_ticketed
            new_sheet["H14"] = f'{i} of {len(cartons)}'
            
        label_wb.remove(template)

        label_wb.save(out_path)
        saved_count += 1
        print("Saved label to:", out_path)

    if saved_count > 0:
        messagebox.showinfo("Done", f"{saved_count} label file(s) saved to:\n\n{destination_folder_path}")
    else:
        messagebox.showinfo("No Files Saved", "No labels were generated due to overwrite selections or errors.")


def generate_template2_labels():
    if not is_valid_path(source_folder_path, destination_folder_path):
        return

    print("Running Template 2 label logic...")
    print("From:", source_folder_path)
    print("To:", destination_folder_path)

    # Grab user inputs
    sync_style_metadata()

    saved_count = 0
    for file in get_input_files(source_folder_path): # Loop xlsx files in destination folder
        if file.name.startswith("~$"):
            print("Skipping temporary file:", file.name) #Skip temporary files created by Excel
            continue

        out_path = Path(destination_folder_path) / f"{file.stem}-LABELS.xlsx"

        if not confirm_overwrite_if_needed(out_path):
            print("Skipped:", out_path.name)
            continue

        print("Processing ", file.name)
    
        # Load input packing list
        source_wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
        packing_list = source_wb.active

        # Load template workbook
        template_path = base_path / "templates" / "template2.xlsx"
        label_wb = openpyxl.load_workbook(template_path, data_only=True)
        template = label_wb.active

        # === Parse and display header info ===
        header = parse_packing_header(packing_list)
        print("Header data:", header)

        # === Parse and display carton info ===
        cartons = parse_packing_list(packing_list)
        print("Carton data: ", cartons)


        for i, carton in enumerate(cartons, start=1):
            print(f"Carton {i} of {len(cartons)}")
            
            new_sheet = label_wb.copy_worksheet(template)
            new_sheet.title = f"Carton {i}"
            ratio, qtys = get_size_ratio_string(carton)

            # TODO: Manually map carton + header values to template cells
            new_sheet["D3"] = header["shipper_address_line1"]
            new_sheet["D4"] = header["shipper_address_line2"]
            new_sheet["D5"] = header["shipper_address_line3"]
            new_sheet["D7"] = header["ship_to_address_line1"]
            new_sheet["D8"] = header["ship_to_address_line2"]
            new_sheet["D9"] = header["ship_to_address_line3"]
            new_sheet["D11"] = header["po_box"]
            #new_sheet["E12"] = header["dept_num"]
            new_sheet["E13"] = carton["vendor_style"]
            new_sheet["E14"] = carton["description"]
            new_sheet["E15"] = ratio
            new_sheet["E16"] = f'{i} of {len(cartons)}'
            new_sheet["E17"] = carton["weight"]
            new_sheet["E18"] = len(cartons)
            
        label_wb.remove(template)

        label_wb.save(out_path)
        saved_count += 1
        print("Saved label to:", out_path)

    if saved_count > 0:
        messagebox.showinfo("Done", f"{saved_count} label file(s) saved to:\n\n{destination_folder_path}")
    else:
        messagebox.showinfo("No Files Saved", "No labels were generated due to overwrite selections or errors.")


def generate_template3_labels():
    if not is_valid_path(source_folder_path, destination_folder_path):
        return


    print("Running Template 3 label logic...")
    print("From:", source_folder_path)
    print("To:", destination_folder_path)

    # Grab user inputs
    sync_style_metadata()

    saved_count = 0
    for file in get_input_files(source_folder_path): # Loop xlsx files in destination folder
        if file.name.startswith("~$"):
            print("Skipping temporary file:", file.name) #Skip temporary files created by Excel
            continue

        out_path = Path(destination_folder_path) / f"{file.stem}-LABELS.xlsx"

        if not confirm_overwrite_if_needed(out_path):
            print("Skipped:", out_path.name)
            continue

        print("Processing ", file.name)

        source_wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
        packing_list = source_wb.active

        template_path = base_path / "templates" / "template3.xlsx"
        label_wb = openpyxl.load_workbook(template_path, data_only=True)
        template = label_wb.active

        # Remove all extra sheets except the template
        for sheet in label_wb.worksheets:
            if sheet != template:
                label_wb.remove(sheet)

        header = parse_packing_header(packing_list)
        print("Header data:", header)

        cartons = parse_packing_list(packing_list)
        print("Carton data:", cartons)

        for i, carton in enumerate(cartons, start=1):
            print(f"Carton {i} of {len(cartons)}")

            key = (file.name, carton["vendor_style"], carton["description"])
            meta = style_metadata.get(key)
            color = meta["color"] if meta else template3_color_var.get().strip()

            style = meta["template3_style"] if meta else template3_style_var.get().strip()

            new_sheet = label_wb.copy_worksheet(template)
            new_sheet.title = f"Carton {i}"
            ratio, qtys = get_size_ratio_string(carton)

            # Mappings for template3
            new_sheet["D2"] = header["ship_to_address_line1"]
            new_sheet["D3"] = header["ship_to_address_line2"]
            new_sheet["D4"] = header["ship_to_address_line3"]
            new_sheet["D5"] = header["ship_to_address_line4"]
            new_sheet["D6"] = header["po_box"]
            new_sheet["D7"] = style
            new_sheet["D8"] = carton["description"]
            new_sheet["D9"] = color

            for j, qty in enumerate(carton["size_quantities"][1:]):
                new_sheet.cell(row=12, column=j + 4).value = qty

            new_sheet["D13"] = carton["weight"]
            new_sheet["D14"] = carton["carton_dimension1"]
            new_sheet["F14"] = carton["carton_dimension2"]
            new_sheet["H14"] = carton["carton_dimension3"]
            new_sheet["D15"] = i
            new_sheet["F15"] = len(cartons)

        label_wb.remove(template)

        label_wb.save(out_path)
        saved_count += 1
        print("Saved label to:", out_path)

    if saved_count > 0:
        messagebox.showinfo("Done", f"{saved_count} label file(s) saved to:\n\n{destination_folder_path}")
    else:
        messagebox.showinfo("No Files Saved", "No labels were generated due to overwrite selections or errors.")


# ============ UI CREATION ===============

# === Folder Selection ===
tk.Checkbutton(window, text="Select single file instead", variable=source_mode, onvalue="file", offvalue="folder").pack()
tk.Button(window, text="Choose Source", command=choose_source).pack()
source_label = tk.Label(window, text="No source folder selected")
source_label.pack()

tk.Button(window, text="Select Destination Folder", command=choose_destination_folder).pack(pady=(20, 5))
destination_label = tk.Label(window, text="No destination folder selected")
destination_label.pack(pady=2)

# === Dynamic UI Frame for Styles ===
style_metadata = {}
style_fields = {}  # To track text input widgets per (style, desc)
auto_style_var = tk.BooleanVar()

def on_auto_style_toggle():
    update_style_fields()
    toggle_template_inputs()


def on_template_change(event):
    template1_frame.pack_forget()
    template3_frame.pack_forget()
    update_style_fields()
    selected = template_var.get()
    if selected == "Template 1":
        template1_frame.pack(pady=(10, 5))
    elif selected == "Template 3":
        template3_frame.pack(pady=(10, 5))
    update_auto_style_visibility()

def update_style_fields():
    for widget in style_inner_frame.winfo_children():
        widget.destroy()
    style_metadata.clear()
    style_fields.clear()

    if not auto_style_var.get():
        style_frame.pack_forget()
        return

    grouped_styles = collect_unique_styles()
    current_template = template_var.get()

    for filename, style_list in grouped_styles.items():
        # Section label per file
        tk.Label(style_inner_frame, text=f"📦 File: {filename}", font=("Arial", 10, "bold")).pack(anchor="w", pady=(10, 2))

        for i, (style, desc) in enumerate(style_list, 1):
            tk.Label(style_inner_frame, text=f"{i}. {style} — {desc}").pack(anchor="w", pady=(5, 0))

            entry_row = {}

            color_var = tk.StringVar()
            tk.Label(style_inner_frame, text="Color:").pack(anchor="w")
            tk.Entry(style_inner_frame, textvariable=color_var).pack(fill="x")
            entry_row["color"] = color_var

            if current_template == "Template 3":
                template3_var = tk.StringVar()
                tk.Label(style_inner_frame, text="Style:").pack(anchor="w")
                tk.Entry(style_inner_frame, textvariable=template3_var).pack(fill="x")
                entry_row["template3_style"] = template3_var

            # 🔑 Now use (filename, style, desc) as the key
            style_fields[(filename, style, desc)] = entry_row

    style_frame.pack(fill="both", expand=True, padx=10, pady=(10, 5))

def toggle_template_inputs():
    is_auto = auto_style_var.get()

    # Hide/show inputs in Template 1
    for widget in template1_frame.winfo_children():
        if isinstance(widget, tk.Entry) or (isinstance(widget, tk.Label) and "Color" in widget.cget("text")):
            widget.pack_forget() if is_auto else widget.pack(fill="x", pady=(0, 10))

    # Hide/show inputs in Template 3
    for widget in template3_frame.winfo_children():
        if isinstance(widget, tk.Entry) or (isinstance(widget, tk.Label) and ("Color" in widget.cget("text") or "Style" in widget.cget("text"))):
            widget.pack_forget() if is_auto else widget.pack(fill="x", pady=(0, 10))

# === Template Selection ===
tk.Label(window, text="Choose Template:").pack(pady=(15, 2))
template_dropdown = ttk.Combobox(window, textvariable=template_var, values=["Template 1", "Template 2", "Template 3"], state="readonly")
template_dropdown.pack(pady=2)

# === Template-Specific Dynamic Frames ===
# Frame for Template 1
template1_frame = tk.Frame(window)
tk.Checkbutton(template1_frame, text="Store Ready", variable=store_ready_var).pack(anchor="w")
tk.Checkbutton(template1_frame, text="Pre-Ticketed", variable=pre_ticketed_var).pack(anchor="w")
tk.Label(template1_frame, text="Color:").pack(anchor="w")
tk.Entry(template1_frame, textvariable=template1_color_var).pack(fill="x", pady=(0, 10))
template1_frame.pack_forget()

# Frame for Template 1
template3_frame = tk.Frame(window)
tk.Label(template3_frame, text="Color:").pack(anchor="w")
tk.Entry(template3_frame, textvariable=template3_color_var).pack(fill="x", pady=(0, 10))
tk.Label(template3_frame, text="Style:").pack(anchor="w")
tk.Entry(template3_frame, textvariable=template3_style_var).pack(fill="x", pady=(0, 10))
template3_frame.pack_forget()

template_dropdown.bind("<<ComboboxSelected>>", on_template_change)

# Checkmark for multiple style inputs
auto_style_frame = tk.Frame(window)
auto_style_checkbox = tk.Checkbutton(auto_style_frame, text="Input multiple styles", variable=auto_style_var, command=on_auto_style_toggle)
auto_style_checkbox.pack()
auto_style_frame.pack_forget() # hide initially

tk.Button(window, text="Generate Labels", command=generate_labels).pack(pady=(30, 5))



style_frame = tk.LabelFrame(window, text="Product Field Overrides")
style_canvas = tk.Canvas(style_frame, height=200)
scrollbar = tk.Scrollbar(style_frame, orient="vertical", command=style_canvas.yview)
style_inner_frame = tk.Frame(style_canvas)
style_inner_frame.bind(
    "<Configure>", lambda e: style_canvas.configure(scrollregion=style_canvas.bbox("all"))
)
style_window_id = style_canvas.create_window((0, 0), window=style_inner_frame, anchor="nw")
style_canvas.configure(yscrollcommand=scrollbar.set)
style_canvas.pack(side="left", fill="both", expand=True)
scrollbar.pack(side="right", fill="y")
style_frame.pack_forget()  # Hide until activated



def bind_mousewheel(widget, canvas):
    def _on_mousewheel(event):
        canvas.yview_scroll(-1 * (event.delta // 120), "units")

    widget.bind_all("<MouseWheel>", _on_mousewheel)  # For Windows
    widget.bind_all("<Button-4>", lambda e: canvas.yview_scroll(-1, "units"))  # For Linux
    widget.bind_all("<Button-5>", lambda e: canvas.yview_scroll(1, "units"))   # For Linux

def resize_inner_frame(event):
    canvas_width = event.width
    style_canvas.itemconfig(style_window_id, width=canvas_width)

bind_mousewheel(style_inner_frame, style_canvas)
style_canvas.bind("<Configure>", resize_inner_frame)


# === Start GUI ===
window.mainloop()
